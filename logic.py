import json
import re
import random

import requests
from requests.adapters import HTTPAdapter

import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.writer.excel import save_workbook


TIME_OUT = 30
MAX_REDIRECT = 50


async def get_proxy():
    """Получение прокси из файла"""

    # Получаем содержимое файла
    with open('proxy.json', 'r', encoding='utf-8') as f:
        proxi = json.load(f)['ok']

    return random.choice(proxi)


def clean_out_excel():
    """Если есть выходной файл Excel удаляет его и создает новый"""
    if os.path.exists('Reels.xlsx') is True:
        os.remove('Reels.xlsx')


def read_setup() -> tuple:
    """Получает данные для парсинга"""
    with open('users_to_pars.txt', 'r', encoding='utf-8') as file:
        users = file.read().split('\n')

    view = int(users[0])
    users = users[1:]
    return view, users


def creat_out_excel():
    """Создает новый выходной Excel файл"""
    file_name = 'Reels.xlsx'
    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet'])
    ws = wb.create_sheet('reels')
    ws.cell(row=1, column=1).value = 'ссылка'
    ws.cell(row=1, column=2).value = 'просмотров:'
    ws.cell(row=1, column=3).value = 'лайков:'
    ws.cell(row=1, column=4).value = 'комментов:'
    save_workbook(wb, file_name)


def wright_in_excel(reels, cur):
    """Записывает данные рилсов одного аккаунта"""
    wb = load_workbook('Reels.xlsx')
    ws = wb['reels']
    for reel in reels:
        ws.cell(row=cur, column=1).value = reel['url']
        ws.cell(row=cur, column=2).value = reel['play_count']
        ws.cell(row=cur, column=3).value = reel['like_count']
        ws.cell(row=cur, column=4).value = reel['comment_count']
        cur += 1
    save_workbook(wb, 'Reels.xlsx')
    return cur


def load_work_profile() -> dict:
    """Загружает куки"""
    with open('cookies.json', 'r', encoding='utf-8') as f:
        cookies = json.load(f)['ok']

    return random.choice(cookies)


def load_patterns():
    """Загружает потерны"""
    with open('patterns.json', 'r', encoding='utf-8') as f:
        return json.load(f)


def insert_params_in_data(parameters: dict):
    """Вставляем аргументы в data запроса"""
    patterns = load_patterns()
    data = patterns['data_for_reels']
    for p in ['av', 'rev', '__hsi', 'fb_dtsg', 'jazoest', 'lsd', '__spin_r', '__spin_b', '__spin_t']:
        data[p] = parameters[p]

    data['variables'] = data['variables'].replace('userID', parameters['target_id'])

    return data


def insert_cur(data: dict, cur: str, user_id) -> dict:
    """обновляем курсор бд"""
    data['variables'] = (f'{{"after":"{cur}","before":null,"data":{{"include_feed_video":true,'
                         f'"page_size":12,"target_user_id":"{user_id}"}},"first":4,"last":null}}')
    return data


def data_headers(res, q_count):
    """Обработчик данных рилсов из ответа"""

    raw_data = res.json()['data']['xdt_api__v1__clips__user__connection_v2']['edges']
    valid_reels = []

    for n, video in enumerate(raw_data):
        video = video['node']['media']
        if video['play_count'] is not None:
            play_count = video['play_count']
        elif video['view_count'] is not None:
            play_count = video['view_count']
        else:
            play_count = 1

        if play_count >= q_count:
            reels = {'url': f'https://www.instagram.com/reel/{video['code']}',
                     'play_count': play_count}
            if 'like_count' in video:
                reels['like_count'] = video['like_count']
            if 'comment_count' in video:
                reels['comment_count'] = video['comment_count']
            valid_reels.append(reels)

    return valid_reels


def check_end(res):
    """Проверяет последний ли это срез"""
    res = res.json()['data']['xdt_api__v1__clips__user__connection_v2']['page_info']['has_next_page']
    if res:
        return False
    else:
        return True


class ParsAccountReels:
    def __init__(self, account_name: str, q_count: int):

        self.account_name = account_name
        self.q_count = q_count
        self.profile_cookies = load_work_profile()
        self.proxy = load_work_profile()

        self.max_retries = 25
        self.time_out = 30

        self.session = requests.Session()
        self.session.mount('https://', HTTPAdapter(max_retries=self.max_retries))
        self.session.proxies.update(self.proxy)

        self.patterns = load_patterns()
        self.cur = None
        self.reels = []
        self.order = 0

    async def swap_work_profile(self, status: str):
        """Меняет рабочий аккаунт"""

        with open("cookies.json", 'w', encoding='utf-8', ) as f:
            """Открываем файл с аккаунтами"""

            accounts = json.load(f) # Загружаем список рабочих аккаунтов

            # Проверяем статус аккаунта
            if self.profile_cookies in accounts['ok']:
                accounts['ok'].remove(self.profile_cookies) # добавляем в неактивные
                accounts[status].append(self.profile_cookies)
                json.dump(accounts, f) # обновляем файл

        self.profile_cookies = load_work_profile() # Обновляем рабочий аккаунт
        await self.reload_session()

    async def change_proxy(self):
        """Меняет рабочий прокси"""

        with open("proxy.json", 'w', encoding='utf-8', ) as f:
            """Открываем файл с прокси"""
            proxies = json.load(f) # Загружаем список прокси

            # Проверяем не добавлен ли он уже в неактивные
            if self.proxy in proxies['ok']:
                proxies['ok'].remove(self.proxy) # добавляем в неактивные
                proxies['end'].append(self.proxy)  # добавляем в неактивные
                json.dump(proxies, f) # обновляем файл

        self.proxy = get_proxy() # Обновляем рабочий прокси

    async def reload_session(self):
        """Перезагружает сессию"""

        self.session.close() # Закрываем текущею
        self.session = requests.Session() # Создаем новую сессию
        self.session.mount(
            'https://',
            HTTPAdapter(max_retries=self.max_retries)
        ) # Монтируем адаптер
        self.session.proxies.update(self.proxy) # Включаем прокси

    async def insert_params_in_headers(self, parameters: dict, referer) -> dict:
        """Вставляем аргументы в headers запроса"""
        patterns = self.patterns
        cookies = self.profile_cookies
        headers_for_reels = patterns['headers_for_reels']
        headers_for_reels['referer'] = referer
        headers_for_reels['x-bloks-version-id'] = parameters['x_bloks_version_id']
        headers_for_reels['x-csrftoken'] = cookies['csrftoken']
        headers_for_reels['x-fb-lsd'] = parameters['lsd']
        headers_for_reels['x-ig-app-id'] = parameters['app_id']
        return headers_for_reels

    async def get_base_html(self):
        """Получаем базовый html аккаунта, для дальнейших запросов"""

        # получаем все заголовки для запроса
        headers = self.patterns['headers_for_html']
        headers['referer'] = headers['referer'].replace('name', self.account_name)

        # запрос
        try:
            base = self.session.get(f'https://www.instagram.com/{self.account_name}/reels/',
                                    cookies=self.profile_cookies, headers=headers, timeout=self.time_out)
        except requests.exceptions.Timeout:
            print('\ntimout')
            return self.get_base_html()
        except requests.exceptions.ConnectionError:
            print('account time ban\n')
            await self.change_proxy()
            await self.swap_work_profile('time_ban')
            return await self.get_base_html()
        except requests.exceptions.TooManyRedirects:
            await self.change_proxy()
            await self.reload_session()
            return await self.get_base_html()


        # проверяем статус ответа
        if base.status_code == 200:
            # возвращаем базовый html
            return base.text
        elif base.status_code in [560, 572]:
            # если рабочий аккаунт заблокирован, меняем его
            print('\nwork account baned')
            await self.change_proxy()
            await self.swap_work_profile('full_ban')
            return self.get_base_html()


    async def param_from_html(self, html) -> dict:
        """Получаем аргументы из html"""
        args = {
            'x_bloks_version_id': r'."versioningID":"(.*?)"',
            'lsd': r'"LSD",.*?,."token":"(.*?)"',
            'app_id': r',"APP_ID":"(.*?)"',
            'av': r'actorID":"(.*?)"',
            'rev': r'"rev":(.*?).,',
            '__hsi': r',"hsi":"(.*?)"',
            'fb_dtsg': r'."DTSGInitialData",..,."token":"(.*?)"',
            'jazoest': r'&jazoest=(.*?)"',
            '__spin_r': r'"__spin_r":(.*?),',
            '__spin_b': r',"__spin_b":"(.*?)",',
            '__spin_t': r',"__spin_t":(.*?),',
            'target_id': r'"target_id":"(.*?)"'
        }

        try:
            for parm in args:
                new = re.search(args[parm], html, flags=re.DOTALL | re.MULTILINE).group(1)
                args[parm] = new

            return args
        except Exception as e:
            print(e)
            await self.swap_work_profile('time_ban')
            await self.reload_session()
            html = await self.get_base_html()
            return await self.param_from_html(html)

    async def first_videos(self, parameters) -> dict:
        headers = await self.insert_params_in_headers(parameters,
                                           self.patterns['headers_for_html']['referer'])
        data = insert_params_in_data(parameters)

        # Делаем запрос к api для получения первых 12ти видео
        try:
            first = self.session.post(
                'https://www.instagram.com/graphql/query',
                cookies=self.profile_cookies, headers=headers, data=data, timeout=self.time_out)
        except requests.exceptions.Timeout:
            print('\ntimout')
            return await self.first_videos(parameters)
        except requests.exceptions.ConnectionError:
            print('account time ban\n')
            await self.change_proxy()
            await self.swap_work_profile('time_ban')
            return await self.first_videos(parameters)

        # Проверяем статус запроса
        if first.status_code == 200:
            print(f'Получено: {12} видео', end='\r')
        elif first.status_code in [560, 572]:
            # если рабочий аккаунт заблокирован, меняем его
            print('\nwork account baned')
            await self.swap_work_profile('full_ban')
            return await self.first_videos(parameters)
        else:
            return {'ok': False, 'error': first.status_code}

        try:
            first.json()
            return {'ok': True, 'res': first}
        except requests.exceptions.JSONDecodeError:
            print('\naccount time ban')
            await self.swap_work_profile('time_ban')
            return await self.first_videos(parameters)


    async def subsequent_videos(self, parameters, cur) -> dict:
        data = insert_params_in_data(parameters)
        data = insert_cur(data, cur, parameters['target_id'])
        headers = await self.insert_params_in_headers(parameters,
                                           self.patterns['headers_for_html']['referer'])

        # Делаем запрос для получения следующих 12ти видео
        try:
            response = self.session.post(
                'https://www.instagram.com/graphql/query',
                cookies=self.profile_cookies, headers=headers, data=data, timeout=self.time_out)
        except requests.exceptions.Timeout:
            print('\ntimout')
            return await self.subsequent_videos(parameters, cur)

        except requests.exceptions.ConnectionError:
            print('account time ban\n')
            await self.change_proxy()
            await self.swap_work_profile('time_ban')
            response = self.subsequent_videos(parameters, data)

        # Проверяем статус запроса
        if response.status_code == 200:
            try:
                response.json()
            except requests.exceptions.JSONDecodeError:
                print(f'всего получено: {self.order * 12} видео, валидных: {len(self.reels)}')
                return {'ok': True, 'next': False, 'data': self.reels}
            self.order += 1
            print(f'Получено: {self.order * 12} видео', end='\r')

            if 'errors' in response.json():
                print(f'всего получено: {self.order * 12} видео, валидных: {len(self.reels)}')
                return {'ok': True, 'next': False, 'data': self.reels}

            elif 'data' in response.json():
                videos = data_headers(response, self.q_count)
                self.reels.extend(videos)
                if check_end(response):
                    print(f'всего получено: {self.order * 12} видео, валидных: {len(self.reels)}')
                    return {'ok': True, 'next': False, 'data': self.reels}

            return {'ok': True, 'res': response, 'next': True}

        else:
            return {'ok': False, 'error': response.status_code, 'next': False}


    async def pars(self) -> dict:
        base_html = await self.get_base_html()

        # Пробуем получить доп параметры для запроса рилсов
        parameters = await self.param_from_html(base_html)

        first_request = await self.first_videos(parameters)

        if first_request['ok']:
            videos = data_headers(first_request['res'], self.q_count)
            self.reels.extend(videos)
        else:
            return {'ok': False, 'error': first_request['error']}

        # Проверяем конец-ли это
        self.order = 1
        if check_end(first_request['res']):
            print(f'всего получено: {self.order * 12} видео, валидных: {len(self.reels)}')
            return {'ok': True, 'data': self.reels}

        self.cur = first_request['res'].json()['data'][
            'xdt_api__v1__clips__user__connection_v2']['page_info']['end_cursor']

        while True:
            # Получаем курсор для следующего запроса
            subsequent_requests = await self.subsequent_videos(parameters, self.cur)
            if subsequent_requests['ok'] and subsequent_requests['next']:
                self.cur = subsequent_requests['res'].json()['data'][
                    'xdt_api__v1__clips__user__connection_v2']['page_info']['end_cursor']
            else:
                return {'ok': True, 'data': self.reels}
